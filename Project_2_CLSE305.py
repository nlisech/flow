# -*- coding: utf-8 -*-
"""
Project 2
Created on Thu Sep 12 22:17:28 2019
# This program was created by Nathan Lisech
# For VCU's CLSE 305 Thermo II

@author: Nathan Lisech
"""
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.chart import (LineChart)
from thermo.chemical import Chemical
import numpy as np
import tkinter

#Rename filepath to location of Excel Worksheet
filepath = r"C:\Users\Nathan Lisech\.spyder-py3\Project 1 CLSE 305.xlsx"
wb = load_workbook(filepath)
worksheet = wb.active

"""
Antoine Equation
Used to find the Psat
 log (P) = A - (B / (T+C) )
"""
T = 100 #degrees C
Tk = T + 273 # degrees K


class Chemical:
    def __init__(self, AntA, AntB, AntC):
        self.AntA = AntA
        self.AntB = AntB
        self.AntC = AntC
        
        
    def Antoine(self):
            Psaturation = (self.AntA - (self.AntB / (Tk + self.AntC) ))
            #print(Psaturation)
            global Psat
            Psat = 10**(Psaturation)
            global PsatRound
            PsatRound = round(Psat, 4)
            print(PsatRound, "Saturated Vapor Pressure in Bar ")
            return PsatRound
            
    def Raoults(X, P):
        for X in range(0, 21):
            global xi
            xi = (X * .05) 
            print() #white space
            print(xi, "X Mole fraction ")
            area = worksheet['B5' : 'K26']
            for row in area:
                for cell in row:
                    worksheet.cell(row=5+X,column=3,value=xi)
            
            for P in range(1,22):
                Pstep = (P * 0.05)
                #print(Pstep, "Pstep in barr")
                print() #white space
                R_Eq = (xi * Psat ) / Pstep
                print(R_Eq)
               # R_EqRound = round(R_Eq, 4)
                #print(R_EqRound, "Y1")
                area = worksheet['B5' : 'K26']
                for row in area:
                    for cell in row:
                        worksheet.cell(row=5+P,column=4,value=Pstep)
                
                area = worksheet['B5' : 'K26']
                for row in area:
                    for cell in row:
                        worksheet.cell(row=5+X,column=2,value=R_Eq)
                        
                plt.title("Pxy")
                plt.ylabel("Pressure in bar")
                plt.xlim(0,1)
                plt.xlabel ('Composition (mole fraction x,y)')
                #plt.legend(loc=1)
                plt.plot(R_Eq, Pstep, 'rs')
            plt.show()
"""
window = tkinter.Tk()
label = tkinter.Label(text="Please choose your chemical.")
label.pack()
l = tkinter.Listbox(window, height=5)
#l.grid(column=0, row=0, sticky=(N,W,E,S))


window.mainloop()


                    #    X1 * Psat + X2 * Psat = (Y1+Y2) P
"""
Heptane =  Chemical(4.02832, 1268.636, -56.199 )
        #AntA = 4.02832
        #AntB = 1268.636
        #AntC = -56.199
        #n-heptane antoine coefficients (299 - 373 K) from NIST
Antoine1 = Chemical.Antoine(Heptane)
Raoult1 = Chemical.Raoults(0,1)  
print("Heptane^^")


Decane = Chemical(4.07857, 1501.268, -78.67 )
        #AntA = 4.07857 
        #AntB = 1501.268
        #AntC = -78.67
        #n-decane antoine coefficients (367.63-338.27 K) from NIST
Antoine2 = Chemical.Antoine(Decane)
Raoult2 = Chemical.Raoults(0,1)
print("Decane^^")
"""

def Save():
    wb.save(filepath)
Save()


#For graph number 2 (n-decane at 100 degrees Celsius)