# -*- coding: utf-8 -*-
"""
# Peng-Robinson Equation of State Solver
# This program was created by Nathan Lisech
# 4/18/19 UPDATED 8/30/19

# This program calculates the compressibility factor, Z, for a compound. 
# From this, the Density, Volume, Hdeparture, Sdeparture, and Gibbs Departure are calculated
# using the Peng-Robinson Equation of State (PREoS) and the data is added to an Excel (.xlsx) file.

#To install Thermo library, use "conda install -c conda-forge thermo"

"""
#Import libraries matplotlib, openpyxl, and Dr. Lewinski's FindZ module file. A new module
# called Chemical was imported from the library, thermo.chemical . This module allows the program to access a
# library of chemical properties, including critical properties, heat capacities, and phase change.
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.chart import (LineChart)
from FindZ import FindDen
from thermo.chemical import Chemical
import numpy as np

#Rename filepath to location of Excel Worksheet Project 3 PREoS Solver1.xlsx
filepath=r"C:\Users\Nathan Lisech\.spyder-py3\Project 3 PREoS Solver1.xlsx"
wb = load_workbook(filepath)
worksheet = wb.active


CompoundName = input('What is the compound name? ')
Chem = Chemical(CompoundName)
print(Chem)
worksheet.title = CompoundName + " Peng-Robinson EoS" #Adds CompoundName to title
worksheet['B7'] = CompoundName  #Inputted Compound Name

#The following block of code places values into the worksheet table
# Parameters for pure component
NoIsoT = int(input('What is the desired number of isotherms? ')) #Enter '5' Isotherms
worksheet['B4']= NoIsoT #Number of inputted isotherms
Tstart = int(x=100) #Starting temperature in Celsius
start = str(Tstart)
worksheet['B2'] = Tstart 
Tstep = int(x=0) #Change in Temperature (K)
worksheet['B3'] = Tstep
Tstop = int(Tstart + (Tstep * NoIsoT)) # Ending Temperature in Celsius
Trange = str(Tstop) #str version of Tstop

P_ref = 1 #1 bar
Pstart = float(x=.001) #Starting Pressure in bar
worksheet['D2'] = Pstart
Pstop = int(x=100) #Ending Pressure in bar
worksheet['D3'] = Pstop
Pstep = int(x=5) #Change in Pressure (bar)
worksheet['D4'] = Pstep
print("Initial Conditions : Tstart = 10 K , Pstart = 10 bar")
#R = float(x=83.14) #R-value = cm3 bar K−1 mol−1 
R = float(x=.0000831447)#R-value = m3 bar K−1 mol−1 
worksheet['B5'] = R

#Pulls chemical property information and places into excel file appropriately
Self_Tc = Chem.Tc # T in K
worksheet['B8'] = Self_Tc 
Tc = float(Self_Tc)
print(Tc)

Self_Pc = Chem.Pc #pulls Pc value in Pa
Pc = Self_Pc / (100000) #Converts from units of Pa to bar
print(Pc)
worksheet['B9'] = Pc

self_acentric = Chem.omega # from reference
worksheet['B10'] = self_acentric
acentric = float(self_acentric)
print(acentric)

Self_MW = Chem.MW #MW in g/mol, from reference
worksheet['B11'] = Self_MW 
MW = float(Self_MW) 
print(MW)

Self_Zc = Chem.Zc #Compressibility factor, Z = PV / NRT
worksheet['B12'] = Self_Zc 
Zc = float(Self_Zc)

plt.figure()

"""
#graph trial
area = worksheet['A15' : 'M15']
c2 = LineChart()
c2.title = "PengRobinson"
c2.style = 12
c2.y_axis.title = "Pressure"
c2.x_axis.majorUnit = "(g/cm3)"
c2.x_axis.title = "Density"
worksheet.add_chart(c2, "F1")
"""

#Nested temperature and pressure loop
def PengRobinson(t,p):  
    for t in range(0,NoIsoT): #(min value, max value, step) Temperature in K
        for p in range(0,10): #(min value, max value, step) Pressure in barr
            Tloop = float(Tstart + (Tstep)*t)
            TloopK = Tloop + 273
            TloopKStr = str(TloopK)
            print(TloopK, "Temperature in K") #Converts Temperature from Celsius to K 
            #Places Tloop values into excel file
            area = worksheet['A16' : 'A26']
            for row in area:
                for cell in area:
                    worksheet.cell(row=16,column=1+7*(t),value=(TloopK))
            
            Ploop = float(Pstart + (Pstep)*p)
            RoundedPloop = round(Ploop, 1)
            PloopStr = str(RoundedPloop)
            print(Ploop, "Pressure in bar")
            #Places Ploop values into excel file
            area = worksheet['B16' : 'B25']
            for row in area: 
                for cell in area:
                    worksheet.cell(row=16+p,column=2+7*(t),value=(RoundedPloop))
                    
            print("The chemical's phase is :", Chem.phase) #Chemical Phase at specified conditions
            Pr = (Ploop) / (Pc) #Reduced Pressure
            RoundedPr = round(Pr, 5)
            print(RoundedPr, "Pr ")
            Tr = (TloopK) / (Tc) #Reduced Temperature 
            RoundedTr = round(Tr, 5)
            print(RoundedTr, "Tr ")
            
            #Calculates slope to solve for alpha
            K = (0.37464 + 1.54226*acentric - (0.26992*acentric*acentric))

           #print("The slope of the line is:", K)
            alpha = (1 + (K*(1-(np.sqrt(Tr)))))
            Roundedalpha = round(alpha, 5)
            print(Roundedalpha, "alpha")
            if Chem.phase == 'g':
                NNN = 0
            if Chem.phase == 'l':
                NNN = 1
                
            A = (0.45724   * ((R**2 * Tc**2)  / Pc)) * alpha**2
           #A = 2.912736 * (10**-6) #from Mixing Rules PREOS Program
            print(A, "A")
            B = 0.07780 *  ((R*Tc) / Pc)
           # B = 3.2498 * (10**-5) #from Mixing Rules PREOS Program
            print(B, "B")
            AA = (A  * (Ploop)) / ( (R * TloopK )**2)
            BB = (B * (Ploop)) / ( R *(TloopK))
            #print(AA, "AA")
            print(BB, "BB")
            RA1 = 1
            #print(RA1, "RA1")
            RA2 = (BB - 1)
            #print(RA2, "RA2")
            RA3 = (AA) - (2*(BB)) - (3*(BB)**2)
            #print(RA3, "RA3")
            RA4 = (BB)**3 + (BB)**2 - ((AA)*(BB))
            #print(RA4, "RA4")
           
            #Density (g/cm3) = P * MW / R*T*Z 
            Z = FindDen(RA1,RA2,RA3,RA4,NNN)
            print(Z, " Calculated Z value")
            
            rho = ( ((Ploop) * (MW)) / ( R * TloopK * Z *1000))
            volume = ( R * TloopK * Z ) / (Ploop)
            
            print() #white space in answer
            
            #print(volume, "Volume (m3/kmol)")
            RoundedVolume = round(volume, 6)
            print(RoundedVolume, "Volume (m3/mol)")

            #print(rho, "Density (kg/m3)") 
            RoundedRho = round(rho, 3)
            print(RoundedRho, "Density (kg/m3)")
            
            ln = np.log( (Z + 2.414*BB) / (Z - 0.414*BB))
            #print(ln, "ln")
            
            Hdep = R * 100000 * Tc * (Tr * (Z-1) - (2.078 * (1+K)) * (alpha)*(ln))
            RoundedHdep = round(Hdep, 2)
            #print(Hdep, "H departure Function (J/mol)" )
            print(RoundedHdep, "H departure Function (J/mol)")           
            #print(Z-BB, "Z-BB")
            
            ln2= np.log( (Z-BB) )
           # print(ln2, "ln2")
            
            Sdep =  R * 100000 * (ln2 -(2.078 * K) * ((1 + K) / (np.sqrt(Tr)) - K) * ln)
            #print(Sdep, "S departure Function  (J/mol*K)")
        
            RoundedSdep = round(Sdep, 4)
            print(RoundedSdep, "S departure Function (J/mol*K)")
            
            IntCp = Chem.HeatCapacityGas.T_dependent_property_integral(298, TloopK)
            #print(IntCp, "Heat Capacity (J/mol)" )
            
            H = (Hdep + IntCp)
            print(H, "H (J/mol)")
            
            IntCpT = Chem.HeatCapacityGas.T_dependent_property_integral_over_T(298, TloopK)
            #print(IntCpT, "Heat capacity over T ( J/mol*K) ") 
            ln3 = np.log(Ploop/P_ref)
            #print(ln3, "ln3")
            #print(R*ln3, " R*ln3")
            S = (-Sdep + IntCpT - R*ln3)
            print(S, "S (J/mol*K)")
            
            AoverB = (AA / (BB * np.sqrt(8)))
            #print(AoverB)
            Gdep = ( (Z-1) - ln2 - (AoverB * ln ))
            print(Gdep, "Gibbs Departure Function (J)")
            fugacity = np.exp(Gdep)
            #print(fugacity, "fugacity") 
            RoundedFug = round(fugacity, 4)
            print(RoundedFug, "Fugacity")
            print() #White space
            
            #Places density values into excel file
            #Places CompoundName into Density header in excel file
            for row in area:
                for cell in area:
                    worksheet.cell(row=15,column=3+7*(t),value=("Density (kg/m^3)"))
            for row in worksheet.iter_rows(min_row=16, max_col=2, max_row=25):
                for cell in row:
                    worksheet.cell(row=16+p,column=3+7*(t),value=RoundedRho)
           #volume values
            for row in area:
                for cell in area:
                    worksheet.cell(row=15,column=4+7*(t),value=("Volume (m^3/mol)"))
            for row in worksheet.iter_rows(min_row=16, max_col=3, max_row=25):
                for cell in row:
                    worksheet.cell(row=16+p,column=4+7*(t),value=RoundedVolume)
            #Hdep values
            for row in area:
                for cell in area:
                    worksheet.cell(row=15,column=5+7*(t),value=("H departure Function (J/mol)"))
            for row in worksheet.iter_rows(min_row=16, max_col=4, max_row=25):
                for cell in row:
                    worksheet.cell(row=16+p,column=5+7*(t),value=RoundedHdep)
            #Sdep values
            for row in area:
                for cell in area:
                    worksheet.cell(row=15,column=6+7*(t),value=("S departure Function (J/mol*K)"))
            for row in worksheet.iter_rows(min_row=16, max_col=5, max_row=25):
                for cell in row:
                    worksheet.cell(row=16+p,column=6+7*(t),value=RoundedSdep)
            #Fugacity values
            for row in area:
                for cell in area:
                    worksheet.cell(row=15,column=7+7*(t),value=("FugacityCoef"))
            for row in worksheet.iter_rows(min_row=16, max_col=6, max_row=25):
                for cell in row:
                    worksheet.cell(row=16+p,column=7+7*(t),value=RoundedFug)
            
            #Plots Density vs Temperature as loop iterates
            plt.title(CompoundName + " Peng-Robinson EoS " + start + " To " + Trange + " deg C")
            plt.ylabel("Pressure in bar")
            plt.xlabel ('Density (kg/m3)')
            #plt.legend(loc=1)
            plt.plot(RoundedRho, Ploop,'rs',label=PloopStr)
            

PengRobinson(0,0)
plt.show()

#Separate Save function, copies worksheet and renames it Compound Name PREoS Starting Temp To Ending Temp
def Save():
    ws1 = wb.copy_worksheet(wb.active)
    ws1.title = (CompoundName + " PREoS " + start + " To " + Trange + " deg C")
    wb.save(filepath)
Save()